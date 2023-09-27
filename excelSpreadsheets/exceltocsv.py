#! python3
# exceltocsv.py - converts excel files to csv file by cmd line

import openpyxl

import csv
import os
import re

filenameRegex=re.compile(r'(^.*)(\.xlsx$)')

os.makedirs('csvFiles',exist_ok=True)

for excelFile in os.listdir('.'):
    #Skip non-excel files,load the workbook object
    if not excelFile.endswith('.xlsx'):
        continue
    else:
        wb=openpyxl.load_workbook(excelFile)
        for sheetName in wb.sheetnames:
            #look through every sheet in the workbook
            sheet=wb[sheetName]
            #create csv filename from the excel filename and sheet title
            csvFilename=(f'{filenameRegex.search(excelFile).group(1)}_{sheetName}.csv')

            #create a csv writer object for this csv file
            csvFile=open(csvFilename,'w',newline='')
            csvWriter=csv.writer(csvFile)
            #loop through every row in the sheet
            rowData = []  # append each cell to this list
            for rowNum in range(0,sheet.max_row):
                colData = []
                for colNum in range(0,sheet.max_column):
                    colData.append((list(sheet.columns)[colNum][rowNum]).value)
                rowData.append(colData)#append each cell's data to rowData.

            #write the rowData list to the CSV file
            for rowNum in range(sheet.max_row):
                csvWriter.writerow(rowData[rowNum])
            print(csvFilename + ' is ready.')
        csvFile.close()
print("All EXCEL files converted to .csv.")
