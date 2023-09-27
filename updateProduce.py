#! python3
#updateProducePrice- corrects cost in product sales spreadsheet

import openpyxl

wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb['Sheet']

#The product types andtheir updated prices
PRICE_UPDATES={'Garlic':3.07,
               'Celery':1.19,
               'Lemon':1.27}
#Loop through rows and update the prices
for rowNum in range(2,sheet.max_row):
    produceName=sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum,column=2).value=PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')